"""
Dynamic Excel Controller Module:
This module provides functions to perform dynamic CRUD (Create, Read, Update, Delete) 
operations on Excel files using the openpyxl library. The functions allow you to write data to the 
first empty row, read data from a specified sheet, update specific cells, and delete cell values in an Excel workbook.
"""
import openpyxl
from openpyxl import load_workbook

def dynamic_write(file_path: str, data: str) -> None:
    # Try to load the existing workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook['SheetName'] # Replace this with actual sheet name
    except FileNotFoundError:
        print(f"File {file_path} not found. Create a new workbook or change the name.")
        return
    
    # Find the first non-empty row
    row = 1
    while sheet.cell(row=row, column=1).value is not None:
        row += 1

    # Write data to the first empty row
    sheet.cell(row=row, column=1).value = data # Adjust column index as needed

def dynamic_read(file_path: str, sheet_name: str) -> list:
    # Open workbook and select the sheet returning an error if wb does not exist
    try: 
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        print(f"File {file_path} not found. Please check the file path.")
        return []
    
    # Create list for data to append and return through itereation
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        return data
    
def dynamic_update(file_path: str, sheet_name: str, row: int, column: int, new_value) -> None:
    # Open the workbook and select the sheet
    try:
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        print(f"File {file_path} not found. Please check the file path.")
        return 
    
    # Update the specific cell with the new value
    sheet.cell(row=row, column=column).value = new_value

    # Save the workbook after updating
    wb.save(file_path)

def dynamic_delete(file_path: str, sheet_name: str, row: int, column: int) -> None:
    try: 
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        print(f"File {file_path} not found. Please check the file path.")
        return
    
    # Clear the value of the specified cell
    sheet.cell(row=row, column=column).value = None

    # Save the workbook after deleting
    wb.save(file_path)







